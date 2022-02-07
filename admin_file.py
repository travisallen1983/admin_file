import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment


column_names = ['First Name', 'Last Name', 'Title', 'Office MLS ID', 'Office_Compare', 'Office E-mail']

df = pd.read_excel('admin_file.xlsx', names = column_names)

unique_offices = set(df['Office MLS ID'])
office_list = list(df['Office MLS ID'])

date = date.today()
date = date.strftime("%B_%d_%Y")

office_admins_full = []

for unique_office in unique_offices:
    office_admins = []
    list_index = 0
    current_office = ''  
    for office in office_list:
        if office == unique_office:
            current_office = office
            first_name = df['First Name'][list_index]
            last_name = df['Last Name'][list_index]
            title = df['Title'][list_index]
            admin_full_name = first_name + ' ' + last_name + ' - ' + title
            office_admins.append(admin_full_name)           
            office_email = df['Office E-mail'][list_index]
        list_index = list_index + 1
    office_admins = str(office_admins).replace('[', '').replace(']', '').replace("'", '').replace('"','')
    office_admins_full.append([current_office, office_email, office_admins])
    
df = pd.DataFrame(office_admins_full, columns = ['Office MLS ID', 'Office E-mail','Office Admins'])            
df.to_excel('Office_Admins' + '_' + date + '.xlsx', index = False)

workbook = load_workbook(filename='Office_Admins' + '_' + date + '.xlsx')
sheet = workbook.active

start = 'C' + str(sheet.min_row + 1)
end = 'C' + str(sheet.max_row)


for row in range(2, len(sheet['C']) + 1):
    print(sheet['C' + str(row)].value)
    clean = sheet['C' + str(row)].value.replace(', ', '\r\n')
    sheet['C' + str(row)] = clean
    sheet['C' + str(row)].alignment = Alignment(wrap_text=True)
    print(sheet['C' + str(row)].value)
    
workbook.save(filename='Office_Admins' + '_' + date + '.xlsx')

